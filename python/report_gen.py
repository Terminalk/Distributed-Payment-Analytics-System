import pandas as pd
from db_connector import get_connection
import logging
import os
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REPORTS_DIR  = os.path.join(PROJECT_ROOT, 'reports')
LOGS_DIR     = os.path.join(PROJECT_ROOT, 'logs')

os.makedirs(REPORTS_DIR, exist_ok=True)
os.makedirs(LOGS_DIR,    exist_ok=True)

logs_format = "%(asctime)s [%(levelname)s] %(message)s"
log_file    = os.path.join(LOGS_DIR, 'report_gen_logs.log')
logging.basicConfig(level=logging.INFO, format=logs_format, filename=log_file, filemode='a')
logger = logging.getLogger(__name__)

HIGH_RISK_THRESHOLD = 30


def fetch_high_risk_users(conn) -> pd.DataFrame:
    query = """
        SELECT
            u.user_id,
            u.external_id,
            u.user_type,
            u.status       AS user_status,
            u.risk_score,
            a.account_id,
            a.account_number,
            a.currency_code,
            a.balance,
            a.available_balance,
            a.status       AS account_status
        FROM users u
        INNER JOIN accounts a ON u.user_id = a.user_id
        WHERE u.risk_score >= :1
        ORDER BY u.risk_score DESC
    """
    return pd.read_sql(query, conn, params=[HIGH_RISK_THRESHOLD])


def fetch_open_alerts(conn) -> pd.DataFrame:
    query = """
        SELECT
            al.alert_id,
            al.alert_type,
            al.alert_score,
            al.threshold_value,
            al.created_at,
            al.generated_by,
            u.user_id,
            u.external_id,
            u.risk_score,
            a.account_id,
            a.account_number
        FROM aml_alert al
        INNER JOIN users    u ON al.user_id    = u.user_id
        INNER JOIN accounts a ON al.account_id = a.account_id
        WHERE al.status = 'OPEN'
        ORDER BY al.alert_score DESC
    """
    return pd.read_sql(query, conn)


def fetch_daily_summary(conn) -> pd.DataFrame:
    query = """
        SELECT
            dab.account_id,
            a.account_number,
            a.currency_code,
            dab.opening_balance,
            dab.closing_balance,
            dab.total_incoming,
            dab.total_outgoing,
            dab.transaction_count,
            dab.balance_date
        FROM daily_account_balance dab
        INNER JOIN accounts a ON dab.account_id = a.account_id
        WHERE dab.balance_date = TRUNC(SYSDATE)
        ORDER BY dab.total_incoming + dab.total_outgoing DESC
    """
    return pd.read_sql(query, conn)


def style_header(ws):
    header_font  = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    header_fill  = PatternFill('solid', start_color='2F5496')
    header_align = Alignment(horizontal='center', vertical='center')

    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 20


def auto_column_width(ws):
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)


def generate_report() -> str:
    logger.info("report_gen: starting report generation")

    conn = get_connection()
    try:
        df_users  = fetch_high_risk_users(conn)
        df_alerts = fetch_open_alerts(conn)
        df_daily  = fetch_daily_summary(conn)
    finally:
        conn.close()

    filename = os.path.join(REPORTS_DIR, f"dpas_report_{date.today()}.xlsx")

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df_users.to_excel(writer,  sheet_name='High Risk Users', index=False)
        df_alerts.to_excel(writer, sheet_name='Open AML Alerts', index=False)
        df_daily.to_excel(writer,  sheet_name='Daily Summary',   index=False)

    wb = load_workbook(filename)
    for ws in wb.worksheets:
        style_header(ws)
        auto_column_width(ws)
    wb.save(filename)

    logger.info(
        f"report_gen: report saved as {filename} — "
        f"users={len(df_users)}, alerts={len(df_alerts)}, daily={len(df_daily)}")
    return filename


if __name__ == '__main__':
    generate_report()